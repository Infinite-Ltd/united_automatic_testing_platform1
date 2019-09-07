from openpyxl import load_workbook
from openpyxl import *
import time


class Logfile:

    #current_row = 1

    def __init__(self, current_row=1):
        '''
        创建对象的时候，初始化一个
        :param current_row: 当前独到的行数（该系统中读取数据库每四行为一个用例的全部组成）
        '''
        self.current_row = current_row
        self.sheet = None
        self.wb = None
        #wb.save(time.strftime('%Y-%M-%D')+'--test_result.xlsx')

    def write_result(self, row, column, message):
        '''
        将自动化用例运行的结果写入Excel文件中
        :param row: 
        :param column: 
        :param message: 
        :return: 
        '''
        #读取Excel文件将放到系统的全局中，现在只是为了写程序方便
        sheet = self.sheet
        # result_excel = load_workbook(time.strftime('%Y-%M-%D')+'result.xlsx') #这句话联调必然会删除的
        # sheetName = result_excel.get_sheet_names()
        # sheet = result_excel.get_sheet_by_name(sheetName[0])
        cell = sheet.cell(row=row, column=column)
        cell.value = message

    def read_file(self):
        result_excel = Workbook()
        self.wb = result_excel
#        result_excel.worksheets()
        #result_excel = load_workbook(time.strftime('%Y-%M-%D') + 'result.xlsx')  # 这句话联调必然会删除的
        self.sheet = result_excel.active
        #self.sheet = result_excel.get_sheet_by_name(sheetName)

    def init_excel(self):
        '''
        初始化用例执行情况的表头
        :return:
        '''
        self.write_result(1, 1, '用例编号')
        self.write_result(1, 2, '用例名称')
        self.write_result(1, 3, '执行结果')
        for i in range(15):
            self.write_result(1, i+4, 'step'+str(i))

#测试部分------------------------------------------------------------
# logfile = Logfile()
# logfile.read_file()
#
# logfile.write_result(1, 1, '测试一下')
# logfile.wb.save(r'c:\test_auto2.xlsx')